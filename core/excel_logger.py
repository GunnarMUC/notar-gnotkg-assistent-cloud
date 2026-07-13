"""Excel-Logger – Revisionssicheres Traceability-Log pro Rechnung."""

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.models import GeneratedInvoice


def create_audit_log(invoice: GeneratedInvoice) -> bytes:
    """Erstellt ein Excel-Traceability-Log für eine Honorarrechnung.

    Args:
        invoice: Die generierte Rechnung mit allen Positionen.

    Returns:
        Excel-Datei als Bytes.
    """
    wb = Workbook()

    # --- Sheet 1: Übersicht ---
    ws = wb.active
    ws.title = "Übersicht"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # Titel
    ws.merge_cells("A1:D1")
    ws["A1"] = "Honorarrechnung – Traceability-Log"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = f"Rechnungs-ID: {invoice.id}"
    ws["A3"] = f"Erstellt: {invoice.created_at.strftime('%d.%m.%Y %H:%M:%S')}"
    ws["A4"] = f"Notar: {invoice.notary.name} – {invoice.notary.firm_name}"
    ws["A5"] = f"Urkunde: {invoice.original_document}"
    ws["A6"] = f"Fee-Engine: {invoice.fee_engine_version}"

    # Summen
    ws["A8"] = "Summen"
    ws["A8"].font = Font(bold=True)
    ws["A9"] = f"Honorar netto: {sum(p.fee_amount for p in invoice.positions):,.2f} €"
    ws["A10"] = f"Auslagen: {sum(invoice.auslagen.values()):,.2f} €"
    ws["A11"] = f"Nettobetrag: {invoice.total_net:,.2f} €"
    ws["A12"] = f"Umsatzsteuer ({invoice.vat_rate * 100:.0f} %): {invoice.vat_amount:,.2f} €"
    ws["A13"] = f"Gesamtbetrag: {invoice.total_gross:,.2f} €"
    ws["A13"].font = Font(bold=True, color="1F4E79")

    ws.column_dimensions["A"].width = 50

    # --- Sheet 2: Detail ---
    ws2 = wb.create_sheet("Positionen")

    headers = [
        "KV-Nr.",
        "Beschreibung",
        "Geschäftswert (€)",
        "Berechnete Gebühr (€)",
        "Fundstelle im Dokument",
        "Manuell geändert",
        "Änderungsgrund",
        "Berechnungsdetails",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row_idx, pos in enumerate(invoice.positions, 2):
        ws2.cell(row=row_idx, column=1, value=pos.kv_number)
        ws2.cell(row=row_idx, column=2, value=pos.description)
        ws2.cell(row=row_idx, column=3, value=pos.business_value_eur)
        ws2.cell(row=row_idx, column=4, value=pos.fee_amount)
        ws2.cell(row=row_idx, column=5, value=pos.source_reference)
        ws2.cell(row=row_idx, column=6, value="Ja" if pos.was_overridden else "Nein")
        ws2.cell(row=row_idx, column=7, value=pos.override_reason or "")
        ws2.cell(row=row_idx, column=8, value=pos.calculation_details)

        # Manuell geänderte Positionen gelb markieren
        if pos.was_overridden:
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for col in range(1, 9):
                ws2.cell(row=row_idx, column=col).fill = yellow_fill

    column_widths = [8, 35, 16, 16, 28, 14, 20, 25]
    for i, width in enumerate(column_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = width

    # --- Sheet 3: Audit ---
    ws3 = wb.create_sheet("Audit-Log")
    ws3["A1"] = "Timestamp"
    ws3["B1"] = "Aktion"
    ws3["C1"] = "Details"
    for col in range(1, 4):
        cell = ws3.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    ws3.cell(row=2, column=1, value=invoice.created_at.isoformat())
    ws3.cell(row=2, column=2, value="Rechnung erstellt")
    ws3.cell(
        row=2,
        column=3,
        value=f"{len(invoice.positions)} Positionen, "
        f"Gesamt: {invoice.total_gross:,.2f} €, "
        f"Fee-Engine: {invoice.fee_engine_version}",
    )

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 70

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
